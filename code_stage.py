
# -*- coding: utf-8 -*-
"""
Created on Tue Jul  9 07:47:57 2024

@author: Lenovo
"""


#from openpyxl import * //cette methode d'importation  n'importe pas tous les bibliotheque
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import math as m
import numpy as np


input_file_path = "C:\\Users\\Lenovo\\Documents\\K.xlsx"
output_file_path = "C:\\Users\\Lenovo\\Documents\\calcul.xlsx"




#///////////////////////////////////////////////////////définiton des fonctions 
#//////////////////////////////////////////////////////


def verif(df):
    k=1
    for col in df.columns:
        for idx, val in enumerate(df[col]):  #utiliser pour parcourir à la fois les indices et les élements
            if isinstance(val, str) or val*-1>0 :
                k=0
                return k
    return k           
                
def extract_second_dataframe(sheet):
    # Identifier les lignes complètement vides
    is_nan = sheet.isna().all(axis=1)
    
    # Trouver le premier et le deuxième bloc de données séparés par une ligne vide
    blocks = []
    current_block = []
    for idx, row in sheet.iterrows():
        if is_nan[idx]:
            if current_block:
                blocks.append(current_block)
                current_block = []
            # Si deux blocs ont été trouvés, on arrête la recherche
            if len(blocks) == 2:
                break
        else:
            current_block.append(row)
    
    # Ajouter le dernier bloc s'il y a encore des données restantes
    if current_block:
        blocks.append(current_block)
    
    # Vérifier si nous avons trouvé au moins deux blocs de données
    if len(blocks) < 2:
        raise ValueError("Moins de deux dataframes trouvés dans la feuille")


    return blocks

def reg_cum(df):
    for i in range(len(df.columns)):
        for j in range (1,len(df.columns)):
            if j-1>=0 and df.iloc[i,j-1]=='nan':
                df.iloc[i,j-1]=0
                
            df.iloc[i,j]+=df.iloc[i,j-1]
            df.iloc[i,j]=np.round(df.iloc[i,j])
    return df           
     
            
def fact_dev(df): 
    fact_dev=[]
    for col in df.columns[:-1]:
        fact_dev.append(df[str(int(col)+1)].sum() / df[col][:-int(col)].sum())
    return fact_dev    




def multiplication_fact_dev(df,fact_dev):
    for i, col in enumerate(df.columns[1:], start=1):
        for j in range(i+1):
            for k in range(len(df) - (j+1)):
                if pd.isna(df.iloc[-(k+1), i]):
                    df.iloc[-(k+1), i] = m.ceil(fact_dev[i-1] *df.iloc[-(k+1), i-1])
                
    return df 


def charge_cum(df1,df2,df3):
    for i in range(len(df1)):
        for j in range(1,len(df2)+1):
            
               df3.iloc[i,j]=float(df1.iloc[i,j])+float(df2.iloc[i,j])

    # Rendre la colonne 'charge' comme index
    df3.set_index('charge', inplace=True)
    return df3
    
                      
def cash_flow(df):
    cash_flows = []
    j = 0
    k = 0
    while k < len(df.columns)-2:
        k += 1
        j = k
        cash_flow = 0
        
        for i in range(len(df)-1, -1, -1):
            if j < len(df.columns)-1:
                cash_flow += df.iloc[i, j+1] - df.iloc[i, j]
                j += 1
        cash_flows.append(cash_flow)
    return cash_flows



            
                
                
def extraire_diagonals(df):
    # trouver diagonals
    rows, cols = df.shape
    diagonals = [[] for _ in range(cols - 1)]
    for i in range(rows):
        for j in range(1, cols):
            if i + j < cols:
                cell = df.iat[i, i + j]
                value = cell if not pd.isna(cell) else 0
                diagonals[j - 1].append(value)
                
    return diagonals


def export(df,nom):
    with pd.ExcelWriter(output_file_path, engine='openpyxl',mode='a') as writer:
        df.to_excel(writer, sheet_name=nom,index=False)
        # Accéder à la feuille Excel et appliquer le style au header
        workbook = writer.book
        worksheet = workbook[nom]

        # Définir le style pour le header en bleu clair
        fill = PatternFill(start_color='00CCFF', end_color='00CCFF', fill_type='solid')  #le type de remplissage solid:remlpissage plein

        # Appliquer le style à chaque cellule dans la première ligne (header)
        for cell in worksheet[1]:
            cell.fill = fill

        # Sauvegarder les modifications dans le fichier Excel
        workbook.save(output_file_path)



#/////////////////////création de la feuille input
df = pd.read_excel(input_file_path, sheet_name='input', engine='openpyxl')
#Extraire le premiér dataframe
first_data_frame=pd.DataFrame(extract_second_dataframe(df)[0]).reset_index(drop=True)
first_data_frame= first_data_frame.drop(first_data_frame.index[0])
# Extraire le deuxième dataframe
second_dataframe = pd.DataFrame(extract_second_dataframe(df)[1]).reset_index(drop=True)
second_dataframe= second_dataframe.drop(second_dataframe.index[0])

if verif(first_data_frame) and verif(second_dataframe) :

    try:
       
        
        # lire la feuille excel
        df = pd.read_excel(input_file_path, sheet_name='input', engine='openpyxl')
        
        # déterminer le 1ére ligne vide 
        first_table_end = df[df.isnull().all(axis=1)].index[0]
        # Trouver toutes les lignes vides
        empty_rows = df[df.isnull().all(axis=1)].index
        D=len(empty_rows)
        # déterminer le debut de la 2éme dataframe
        second_table_start = df[df.isnull().all(axis=1)].index[1]
        # charger Workbook
        workbook = load_workbook(input_file_path)
        sheet = workbook['input']
    
        # créer la nouvelle feuille input1
        new_sheet = workbook.create_sheet(title='input1')
        
        for row in sheet.iter_rows(min_row=1, max_row=first_table_end+1, values_only=False):
            for cell in row:
                new_cell = new_sheet.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                if cell.has_style:
                    new_cell._style = cell._style
                if cell.hyperlink:
                    new_cell.hyperlink = cell.hyperlink
                if cell.comment:
                    new_cell.comment = cell.comment
          
        # Copier les cellules du deuxième DataFrame
        for row in sheet.iter_rows(min_row=second_table_start + 1, values_only=False):
            for cell in row:
                new_cell = new_sheet.cell(row=cell.row - second_table_start + first_table_end+1, column=cell.col_idx, value=cell.value)
                if cell.has_style:
                    new_cell._style = cell._style
                if cell.hyperlink:
                    new_cell.hyperlink = cell.hyperlink
                if cell.comment:
                    new_cell.comment = cell.comment
                    
        workbook.save(output_file_path)                
    
        
    
    except Exception as e:
        print(f"Error while processing the 'input1' sheet: {e}")
        
        
        
    
    #//////////////////////Création de la feuille REG
    
    
    
    try:
      sheet_name = 'input'
      df_input = pd.read_excel(input_file_path, sheet_name=sheet_name)
    
      # Extraire le deuxième dataframe
      second_dataframe = pd.DataFrame(extract_second_dataframe(df_input)[1]).reset_index(drop=True)
    
      # Exporter le deuxième dataframe dans une nouvelle feuille 'REG'
      with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='a',if_sheet_exists='replace') as writer:
          second_dataframe.to_excel(writer, sheet_name='REG', index=False)
            
        
    except Exception as e:
        print(f"Error while processing the 'REG' sheet: {e}")
        
        
    
        
    #////////////////////////////création de la feuille REG_reglé  
    #/////////////////////////////  
    
    
    
    
    try:
        # Lire le DataFrame
        df_reg = pd.read_excel(output_file_path, sheet_name='REG', engine='openpyxl')
        df_reg.columns=['REGLEMENT']+ ['N']+[f'N+{i}' for i in range(1,len(df_reg.columns)-1)]
        # Supprimer les 2 premières lignes
        df_reg = df_reg.drop(df_reg.index[0])
        df_reg.reset_index(drop=True, inplace=True)
        diagonals=extraire_diagonals(df_reg)
        # Trouver le nombre maximum de diagonales
        max_diagonals = max(len(diagonal) for diagonal in diagonals)
        #Obtenir l'année de la première ligne des données de la colonne 'REGLEMENT'
        annee_debut = int(df_reg['REGLEMENT'].iloc[0])
        
        # Créer un dictionnaire
        data = {'REGLEMENT': [annee_debut + i for i in range(max_diagonals)]}
        for k in range(len(diagonals)):
            data[f'N+{k}'] = [diagonals[k][i] if i < len(diagonals[k]) else None for i in range(max_diagonals)]
        # Créerle DataFrame
        final_df2 = pd.DataFrame(data)
        # Arrondir les valeurs dans le DataFrame
        final_df2 = final_df2.round()
        export(final_df2,'REG_reglé')
        
        
    except Exception as e:
        print(f"Error while processing the 'REG_reglé' sheet: {e}")
        
    
     
        
    #//////////////////////création de la feuille SAP 
    #//////////////////////  
    
    
     
    try:
        workbook = load_workbook(output_file_path)
        sheet = workbook['input']
        new_sheet = workbook.create_sheet(title='SAP')
    
        for row in sheet.iter_rows(min_row=1, max_row=first_table_end+1, values_only=False):
            for cell in row:
                new_cell = new_sheet.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                if cell.has_style:
                    new_cell._style = cell._style
                if cell.hyperlink:
                    new_cell.hyperlink = cell.hyperlink
                if cell.comment:
                    new_cell.comment = cell.comment
    
        workbook.save(output_file_path)
    except Exception as e:
        print(f"Error while processing the 'SAP' sheet: {e}")
          
        
        
        
    #///////////////////////création de la feuille SAP_reglé    
    #///////////////////////////////////// 
    
    
    
    
    try:
        
        
        df = pd.read_excel(output_file_path, sheet_name='SAP', engine='openpyxl')
        # Appliquer les nouveaux noms de colonnes au DataFrame
        df.columns = ['SAP'] + [f'SAP{annee_debut+ i}' for i in range(len(df.columns) - 1)]
        
        # Supprimer les deux premières lignes
        df = df.drop(df.index[0])
        df.reset_index(drop=True, inplace=True)
        
        # Extraire diagonals
        diagonals=extraire_diagonals(df)
        max_diagonals = max(len(diagonal) for diagonal in diagonals)
        #Obtenir l'année de la première ligne des données de la colonne 'REGLEMENT'
        annee_debut = int(df['SAP'].iloc[0])
        
        # Créer un dictionnaire
        data = {'REGLEMENT': [annee_debut + i for i in range(max_diagonals)]}
        for k in range(len(diagonals)):
            data[f'N+{k}'] = [diagonals[k][i] if i < len(diagonals[k]) else None for i in range(max_diagonals)]
        final_df = pd.DataFrame(data)
        final_df.columns=['SAP']+ ['N']+[f'N+{i}' for i in range(1,len(final_df.columns)-1)]
        
        # Write the final DataFrame to a new sheet named 'SAP_d'
        export(final_df,'SAP_reglé')
    except Exception as e:
        print(f"Error while processing the 'SAP_reglé' sheet: {e}")
        
      
        
        
    #///////////////////////////////////////calcul du reglement cumulée
    #/////////////////////////////////////////////////////////////////////
    
    
    
    
    df = pd.read_excel(output_file_path,sheet_name='REG_reglé', engine='openpyxl')
    df.columns=['REGLEMENT']+[f'{i}' for i in range(1,len(df)+1)]
    
    df.set_index(df['REGLEMENT'], inplace=True)
    del df['REGLEMENT']
    df=reg_cum(df)
          
    
     
    #////////////////////////////////calcul du facteur du développement 
    
    
    factors=fact_dev(df)
        
    #/////////////////////////////// Remplir les valeurs manquantes
    
    
    df=multiplication_fact_dev(df,factors)
               
    
    
    # /Ajouter la colonne Tail avec les mêmes valeurs que la colonne N+10
    df['Tail'] = df.iloc[:, -1]
    
    # Sauvegarder le DataFrame dans un fichier Excel
    
    
    with pd.ExcelWriter(output_file_path, engine='openpyxl',mode='a') as writer:
        df.to_excel(writer, sheet_name='reglement_cumulé')
        # Accéder à la feuille Excel et appliquer le style au header
        workbook = writer.book
        worksheet = workbook['reglement_cumulé']
    
        # Définir le style pour le header en bleu clair
        fill = PatternFill(start_color='00CCFF', end_color='00CCFF', fill_type='solid')
    
        # Appliquer le style à chaque cellule dans la première ligne (header)
        for cell in worksheet[1]:
            cell.fill = fill
    
        # Sauvegarder les modifications dans le fichier Excel
        workbook.save(output_file_path) 
    
    
        
        
    
    #////////////////////////////////////////calcul du charge cumulée
    #////////////////////////////////////////////////////////////////
    
    
    
    
    
    # Lire les feuilles Excel en DataFrames
    df = pd.read_excel(output_file_path, sheet_name='reglement_cumulé', engine='openpyxl')
    df.columns = ['REGLEMENT'] + [f'{i}' for i in range(1, len(df.columns)-1)] + ['Tail']
    df_sap = pd.read_excel(output_file_path, sheet_name='SAP_reglé', engine='openpyxl')
    # Initialisation d'un tableau vide avec les mêmes colonnes, index de lignes et en-têtes de lignes
    charges = pd.DataFrame(index=df.index, columns=df.columns)
    charges.iloc[:, 0] = df.iloc[:, 0]
    
    # Renommer la première colonne
    charges = charges.rename(columns={'REGLEMENT': 'charge'})
    charges = charges.drop('Tail', axis=1)
    
    charges=charge_cum(df,df_sap,charges)
    
    # ////////////////////////////////////////////Calcul des facteurs
    #////////////////////////////////////////////
    
    
    
    # Calcul des facteurs
    fact_dev_values=fact_dev(charges)
    
       
    
    # ///////////////////////////////////////////////Remplir les valeurs manquantes
    
    
    
    charges=multiplication_fact_dev(charges,fact_dev_values)
    
    # Ajouter la colonne 'charge' avec les années
    charges.insert(0, 'charge', range(df_sap['SAP'][0],df_sap['SAP'][0] + len(charges)))
    
    # Add the 'Tail' column with the same values as the last column
    charges['Tail'] = charges.iloc[:, -1]
    charges.reset_index(drop=True, inplace=True)
    
    
    # Écrire le DataFrame final dans une nouvelle feuille nommée 'charge_développé'
    
    export(charges,'charge_développé')
    
    #/////////////////////////////////calcul_du_cash_flow_reg
    #/////////////////////////////////////////////////////
    
    
    
    
    # Lire le fichier Excel 
    df = pd.read_excel(output_file_path, sheet_name='reglement_cumulé', engine='openpyxl')
    cash_flow_values=cash_flow(df)
     
    df['CF_REG'] = cash_flow_values    
    
    # Lire le fichier Excel
    df1 = pd.read_excel(output_file_path, sheet_name='charge_développé', engine='openpyxl')
    
    cash_flow_values_charges=cash_flow(charges)
    
    
    
    
    #////////////////////////////////calcul ddu cash flow_charge
    #//////////////////////////////
    
    
    cash_flow_values_charges=cash_flow(df1)
    
    
    df1['CF_charges'] = cash_flow_values_charges 
    
    
    # Concaténer les colonnes de plusieurs DataFrames
    df_concatened = pd.concat([ df['CF_REG'], df1['CF_charges']], axis=1)
    
    # Ajouter un index allant de 'N' à 'N+10'
    df_concatened.index = ['N']+[f'N+{i}' for i in range(1,len(df_concatened))]
    # Charger ou créer le fichier Excel
    try:
        book = load_workbook(output_file_path)
        print("Le fichier Excel existe déjà et sera chargé.")
    except FileNotFoundError:
        book = Workbook()
        book.save(output_file_path)
        book = load_workbook(output_file_path)
        print("Le fichier Excel n'existait pas. Un nouveau fichier a été créé.")
    
    # Sélectionner ou créer la feuille de calcul
    sheet_name = 'cash_flow'
    if sheet_name not in book.sheetnames:
        book.create_sheet(sheet_name)
    sheet = book[sheet_name]
    
    # Supprimer les anciennes données dans la feuille de calcul
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            cell.value = None
    # Ajouter les en-têtes des colonnes
    sheet.append([" "] + list(df_concatened.columns))
    # Ajouter les données du DataFrame à la feuille de calcul
    for idx, row in df_concatened.iterrows():
        sheet.append([idx] + row.tolist())
    # Sauvegarder le fichier Excel
    book.save(output_file_path)
    
    
    
    
    
    #//////////////////////////////////////////////////création de la page output1
    #//////////////////////////////////////////////////
    
    
    
    
    #traitement 
    df_reg_anc = pd.read_excel(output_file_path, sheet_name='REG_reglé', engine='openpyxl')
    df_reg_cum_anc = pd.read_excel(output_file_path, sheet_name='reglement_cumulé', engine='openpyxl')
    df_charge_anc = pd.read_excel(output_file_path, sheet_name='charge_développé', engine='openpyxl')
    df_reg_anc.columns=['REGLEMENT']+ ['N']+[f'N+{i}' for i in range(1,len(df_reg_anc.columns)-1)]
    df_reg_cum_anc.columns=['REG_CUM']+['N']+[f'N+{i}' for i in range(1,len(df_reg_cum_anc.columns)-2)]+['Tail']
    df_charge_anc.columns=['CHARGE']+['N']+[f'N+{i}' for i in range(1,len(df_reg_cum_anc.columns)-2)]+['Tail']
    
    export(df_reg_anc,'REG1')
    
    export(df_reg_cum_anc,'reglement_cumulé1')
    
     
    export(df_charge_anc,'charge_développé1')
    
    
    
    #/////////////////////////////////////////////////exportinnnnnggggggggg
    #/////////////////////////////////////////////////
    
    
    
    # Charger le classeur Excel existant avec openpyxl
    book = load_workbook(output_file_path)
    
    # Charger les données depuis les fichiers Excel
    with pd.ExcelFile(output_file_path) as xls:
        df_sap = pd.read_excel(xls, sheet_name='SAP_reglé')
        df_reg = pd.read_excel(xls, sheet_name='REG1')
        df_reg_cum = pd.read_excel(xls, sheet_name='reglement_cumulé1')
        df_charge_dev = pd.read_excel(xls, sheet_name='charge_développé1')
        df_cash_flow = pd.read_excel(xls, sheet_name='cash_flow')
    
    # Préparer les DataFrames pour l'écriture en ajoutant une ligne vide entre chaque DataFrame
    df_sap[' '] = ''  # Ajouter une colonne vide pour séparer les tableaux visuellement
    df_reg[' '] = ''
    df_reg_cum[' '] = ''
    df_charge_dev[' '] = ''
    df_cash_flow[' ']=''
    # Accéder à la feuille 'output1' dans le classeur existant ou la créer
    if 'output1' in book.sheetnames:
        ws = book['output1']
    else:
        ws = book.create_sheet(title='output1')
    
    # Définir le remplissage de l'en-tête
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    
    # Fonction pour écrire les DataFrames dans la feuille 'output1'
    def write_data_to_sheet(rows, worksheet, start_row):
        for r_idx, row in enumerate(rows, start_row):
            for c_idx, value in enumerate(row, 1):
                cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == start_row:
                    cell.fill = header_fill  # Appliquer le style à l'en-tête
        return r_idx + 2  # Retourner la prochaine ligne de départ, plus une ligne vide
    
    # Écrire les données ligne par ligne avec style pour l'en-tête
    start_row = 1
    start_row = write_data_to_sheet(dataframe_to_rows(df_sap, index=False, header=True), ws, start_row)
    start_row = write_data_to_sheet(dataframe_to_rows(df_reg, index=False, header=True), ws, start_row)
    start_row = write_data_to_sheet(dataframe_to_rows(df_reg_cum, index=False, header=True), ws, start_row)
    
    write_data_to_sheet(dataframe_to_rows(df_charge_dev, index=False, header=True), ws, start_row)
    start_row = write_data_to_sheet(dataframe_to_rows(df_charge_dev, index=False, header=True), ws, start_row)
    write_data_to_sheet(dataframe_to_rows(df_cash_flow, index=False, header=False), ws, start_row)
    # Sauvegarder les modifications dans le fichier Excelnnnn
    book.save(output_file_path)
    
    
    
    
    #/////////////////////////////////////////////supprimer des feuilles excel
    #////////////////////////////////////////////
    
    
    
    # Charger le classeur Excel existant avec openpyxl
    book = load_workbook(output_file_path)
    # Supprimer les feuilles 'input' et 'output' si elles existent
    sheets_to_delete = ['input', 'output','REG','SAP','charge_développé','cash_flow','reglement_cumulé','charge_développé1','cash_flow','reglement_cumulé1','REG1','SAP_reglé','REG_reglé']
    for sheet_name in sheets_to_delete:
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            book.remove(sheet)
    # Sauvegarder les modifications dans le fichier Excel
    book.save(output_file_path)
    
    
    
else:
    print("impossible de manipuler le fichier excel")